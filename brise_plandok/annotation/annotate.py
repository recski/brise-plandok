from brise_plandok.constants import ATTRIBUTE_NORM_MAP, GOLD_COLOR, GOLD_PREFIX
import openpyxl
import argparse
import logging
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, PatternFill
from collections import defaultdict
from brise_plandok.annotation.attributes import ATTR_TO_CAT


class Annotate:
    def parse(self, dataset, template, save):
        sentences = []
        labels = []
        sentence_ids = []

        logging.debug("started the creation of the excel sheet")
        for line in dataset:
            sentence_ids.append(line[0])
            sentences.append(line[1])
            labels_for_row = ",".join(
                [
                    ATTRIBUTE_NORM_MAP[label] if label in ATTRIBUTE_NORM_MAP else label
                    for label in line[2].split(",")
                ]
            )
            labels.append(labels_for_row)

        path = template
        wb_obj = openpyxl.load_workbook(path)
        sheet_obj = wb_obj["Dataset"]
        sheet_obj_labels = wb_obj["Label"]

        attribute_to_attribute_class = defaultdict(str)
        for i in range(ord("F"), ord("T") + 1):
            attrs = []
            for col in sheet_obj_labels[chr(i)]:
                if col.value:
                    attrs.append(col.value)
            for attr in attrs[1:]:
                attribute_to_attribute_class[attr] = attrs[0]

        sheet_obj.insert_rows(idx=3, amount=len(sentences))

        data_val = DataValidation(type="list", formula1="=Attribute")
        sheet_obj.add_data_validation(data_val)

        self._add_sentences(
            sentences,
            labels,
            sheet_obj,
            sentence_ids,
            data_val,
            attribute_to_attribute_class,
        )

        wb_obj.save(save)

    def _add_sentences(
        self,
        sentences,
        labels_csv,
        sheet_obj,
        sentence_ids,
        data_val,
        attribute_to_attribute_class,
    ):
        for i, _ in enumerate(sentences):
            label = labels_csv[i]
            split_labels = label.strip().split(",")
            gold = False
            if len(split_labels) > 0 and split_labels[0] == GOLD_PREFIX:
                gold = True
                split_labels = split_labels[1:]

            row_id = i + 3
            self._set_sentence_data(
                row_id, sentences[i], sheet_obj, sentence_ids[i], gold
            )

            if sentence_ids[i].split("_")[-2] == "0":
                self._add_header(row_id, sheet_obj)
            else:
                self._add_data_validation(data_val, sheet_obj, row_id)

                labels = self._clean_labels(list(set(split_labels)))

                if len(labels) > 0 and labels[0]:
                    self.__set_attributes(
                        sheet_obj,
                        self._get_coordinate("C", row_id),
                        self._get_coordinate("D", row_id),
                        labels[0],
                        attribute_to_attribute_class,
                    )

                if len(labels) > 1:
                    self.__set_attributes(
                        sheet_obj,
                        self._get_coordinate("E", row_id),
                        self._get_coordinate("F", row_id),
                        labels[1],
                        attribute_to_attribute_class,
                    )

                if len(labels) > 2:
                    self.__set_attributes(
                        sheet_obj,
                        self._get_coordinate("G", row_id),
                        self._get_coordinate("H", row_id),
                        labels[2],
                        attribute_to_attribute_class,
                    )

                if len(labels) > 3:
                    self.__set_attributes(
                        sheet_obj,
                        self._get_coordinate("I", row_id),
                        self._get_coordinate("J", row_id),
                        labels[3],
                        attribute_to_attribute_class,
                    )

                if len(labels) > 4:
                    self.__set_attributes(
                        sheet_obj,
                        self._get_coordinate("K", row_id),
                        self._get_coordinate("L", row_id),
                        labels[4],
                        attribute_to_attribute_class,
                    )

    def _clean_labels(self, label_split):
        cleaned_labels = label_split.copy()
        for label in label_split:
            if label and label not in ATTR_TO_CAT:
                logging.warning(
                    f"{label} attribute not in the attribute list, will be skipped!"
                )
                cleaned_labels.remove(label)
        return cleaned_labels

    def _get_coordinate(self, column_letter, row_id):
        return column_letter + str(row_id)

    def _add_data_validation(self, data_val, sheet_obj, row_id):
        data_val.add(sheet_obj[self._get_coordinate("C", row_id)])
        data_val.add(sheet_obj[self._get_coordinate("E", row_id)])
        data_val.add(sheet_obj[self._get_coordinate("G", row_id)])
        data_val.add(sheet_obj[self._get_coordinate("I", row_id)])
        data_val.add(sheet_obj[self._get_coordinate("K", row_id)])

        data_val_subclass_C = DataValidation(
            type="list", formula1="==INDIRECT($C${0})".format(str(row_id))
        )
        sheet_obj.add_data_validation(data_val_subclass_C)
        data_val_subclass_C.add(sheet_obj[self._get_coordinate("D", row_id)])

        data_val_subclass_E = DataValidation(
            type="list", formula1="==INDIRECT($E${0})".format(str(row_id))
        )
        sheet_obj.add_data_validation(data_val_subclass_E)
        data_val_subclass_E.add(sheet_obj[self._get_coordinate("F", row_id)])

        data_val_subclass_G = DataValidation(
            type="list", formula1="==INDIRECT($G${0})".format(str(row_id))
        )
        sheet_obj.add_data_validation(data_val_subclass_G)
        data_val_subclass_G.add(sheet_obj[self._get_coordinate("H", row_id)])

        data_val_subclass_I = DataValidation(
            type="list", formula1="==INDIRECT($I${0})".format(str(row_id))
        )
        sheet_obj.add_data_validation(data_val_subclass_I)
        data_val_subclass_I.add(sheet_obj[self._get_coordinate("J", row_id)])

        data_val_subclass_K = DataValidation(
            type="list", formula1="==INDIRECT($K${0})".format(str(row_id))
        )
        sheet_obj.add_data_validation(data_val_subclass_K)
        data_val_subclass_K.add(sheet_obj[self._get_coordinate("L", row_id)])

    def _set_sentence_data(self, row_id, sentence, sheet_obj, sentence_id, gold):
        sentence_id_coord = "A" + str(row_id)
        sentence_coord = "B" + str(row_id)
        sheet_obj[sentence_coord] = sentence
        sheet_obj[sentence_id_coord] = sentence_id
        sheet_obj[sentence_coord].alignment = Alignment(wrapText=True)
        sheet_obj[sentence_coord].font = Font(size=12)
        sheet_obj[sentence_id_coord].alignment = Alignment(wrapText=True)
        sheet_obj[sentence_id_coord].font = Font(size=12)
        if gold:
            sheet_obj[sentence_coord].fill = PatternFill(
                fgColor=GOLD_COLOR, fill_type="solid"
            )
            sheet_obj[sentence_id_coord].fill = PatternFill(
                fgColor=GOLD_COLOR, fill_type="solid"
            )

    def _add_header(self, row_id, sheet_obj):
        for j in range(ord("C"), ord("P") + 1):
            cell_id = chr(j) + str(row_id)
            sheet_obj[cell_id] = "DON'T ANNOTATE THIS SENTENCE"
            sheet_obj[cell_id].fill = PatternFill(fgColor="878787", fill_type="solid")
            sheet_obj[cell_id].alignment = Alignment(
                wrapText=True, horizontal="center", vertical="center"
            )
            sheet_obj[cell_id].font = Font(bold=True)

    def __set_attributes(
        self,
        sheet_obj,
        category_coordinate,
        label_coordinate,
        label,
        attribute_to_attribute_class,
    ):
        sheet_obj[category_coordinate] = attribute_to_attribute_class[label]
        sheet_obj[label_coordinate] = label


def main():
    logging.basicConfig(
        level=logging.WARNING,
        format="%(asctime)s : "
        + "%(module)s (%(lineno)s) - %(levelname)s - %(message)s",
    )
    annotate = Annotate()
    parser = argparse.ArgumentParser(description="Create excel files for annotation.")
    parser.add_argument(
        "--dataset", required=True, help="the path to the dataset to be annotated"
    )
    parser.add_argument(
        "--template",
        required=True,
        metavar="S",
        help="the initial excel to fill with the dataset",
    )
    parser.add_argument(
        "--save", default="annotate.xlsx", metavar="S", help="where to save the excel"
    )

    args = parser.parse_args()

    dataset = []
    with open(args.dataset) as f:
        for line in f:
            line = line.split("\t")
            sentence = line[1]
            sentence_id = line[0]
            label = line[2]
            dataset.append((sentence_id, sentence, label))

    annotate.parse(args.dataset, args.template, args.save)


if __name__ == "__main__":
    main()
