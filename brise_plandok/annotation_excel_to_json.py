import json
import os

import openpyxl
from brise_plandok.convert import Converter

DATASET_SHEET_NAME = "Dataset"
FIRST_DATA_ROW = 3
ATTRIBUTE_OFFSET = 2
ATTRIBUTE_STEP = 2
LABEL_OFFSET = 1

ID_COL = 1
TEXT_COL = 2

ATTRIBUTES_TO_IGNORE = "DON'T ANNOTATE THIS SENTENCE"

JSON_ATTR = "/home/eszter/ownCloud/Shared/BRISE/data/plandok/ana/2021_09/json_attr"
DATA = "/home/eszter/research/brise-nlp/annotation/2021_09/full_data"

GOLD_COLOR = "00FFD700"

def gen_sens_from_xlsx(fn):
    wb = openpyxl.load_workbook(fn)
    sheet = wb[DATASET_SHEET_NAME]
    doc_id = os.path.basename(fn).split(".")[0]
    doc = {
        "id": doc_id,
        "sens": {},
    }
    for id, sen in _gen_sens(sheet, _get_json_attr(doc_id)):
        assert id not in doc["sens"]
        doc["sens"][id] = sen
    _save(doc_id, doc)


def _get_json_attr(doc_id):
    with open(os.path.join(JSON_ATTR, doc_id+".jsonl")) as f:
        return json.load(f)


def _gen_sens(sheet, json_attr):
    for row in range(FIRST_DATA_ROW, sheet.max_row):
        id = sheet.cell(row=row, column=ID_COL).value
        text = sheet.cell(row=row, column=TEXT_COL).value
        gen_attributes = _get_gen_attributes(json_attr, id)
        already_gold_on_annotation = _is_already_gold(sheet, row)
        yield id, {
            "id": id,
            "text": text,
            "modality": None,
            "is_gold": False,
            "already_gold_on_annotation": already_gold_on_annotation,
            "gen_attributes_on_annotation": gen_attributes,
            "gen_attributes": [],
            "annotated_attributes": [],
            "gold_attributes": [],
        }


def _get_gen_attributes(doc, id):
    for section in doc["sections"]:
        for sen in section["sens"]:
            if sen["sen_id"] == id:
                return sen["gen_attributes"]


def _is_already_gold(sheet, row):
    cell = sheet.cell(row=row, column=ID_COL)
    return cell.fill.fgColor.rgb == GOLD_COLOR


def _save(doc_id, doc):
    data_file = os.path.join(DATA, doc_id + ".json")
    with open(data_file, "w") as f:
        json.dump(doc, f)


def main():
    fn = "/home/eszter/ownCloud/Shared/BRISE/data/plandok/ana/2021_09/xlsx/6770.xlsx"
    gen_sens_from_xlsx(fn)


if __name__ == "__main__":
    main()
