import argparse
import logging
import os

import openpyxl

from brise_plandok.annotation_process.utils.annotation_converter import (
    AnnotationConverter,
)
from brise_plandok.annotation_process.utils.constants import (
    FullAnnotationExcelConstants,
    FullReviewExcelConstants,
)
from brise_plandok.annotation_process.utils.full_review_excel_generator import (
    FullReviewExcelGenerator,
)
from brise_plandok.attrs_from_gold import SenToAttrMap, full_attrs_from_gold_sen
from brise_plandok.constants import (
    ANNOTATOR_NAME_INDEX,
    EMPTY,
    AnnotatedAttributeFields,
    AttributeFields,
    AttributeTypes,
    DocumentFields,
    FullAnnotatedAttributeFields,
    SenFields,
)
from brise_plandok.utils import dump_json, load_json


class FullAnnotationConverter(AnnotationConverter):
    def __init__(self, args):
        super().__init__(args)
        self.sen_to_attr = SenToAttrMap(args.gold_folder, fuzzy=False, full=True)

    def convert(self, annotated_xlsx_files, output_file, data_file):
        assert data_file is not None
        doc = load_json(data_file)
        self._fill_reviewers(doc, output_file, DocumentFields.FULL_REVIEWERS)
        self._fill_annotated_attributes(annotated_xlsx_files, doc)
        self._fill_with_gold(doc)
        dump_json(doc, data_file)
        self._generate_review_excel(doc, output_file)

    def _fill_annotated_attributes(self, annotated_xlsx_files, doc):
        self._clear_previous_annotation_info(
            doc, DocumentFields.FULL_ANNOTATORS, SenFields.FULL_ANNOTATED_ATTRIBUTES
        )
        for annotated_xlsx in annotated_xlsx_files:
            annotator = os.path.normpath(annotated_xlsx).split(os.path.sep)[ANNOTATOR_NAME_INDEX]
            self._add_annotator(doc, annotator, DocumentFields.FULL_ANNOTATORS)
            self._fill_full_annotations(annotated_xlsx, doc, annotator)

    def _fill_full_annotations(self, annotation_xlsx, doc, annotator):
        workbook = openpyxl.load_workbook(annotation_xlsx)
        ann_sheet = workbook[FullAnnotationExcelConstants.MAIN_SHEET_NAME]

        for row_id in range(FullAnnotationExcelConstants.FIRST_DATA_ROW, ann_sheet.max_row + 1):
            sen_id = ann_sheet.cell(
                row=row_id, column=FullAnnotationExcelConstants.SEN_ID_COL
            ).value
            modality = ann_sheet.cell(
                row=row_id, column=FullAnnotationExcelConstants.MODALITY_COL
            ).value
            self._fill_modality(doc, sen_id, modality, annotator)
            sen = doc[DocumentFields.SENS][sen_id]
            for col in range(
                FullAnnotationExcelConstants.ATTRIBUTE_OFFSET,
                ann_sheet.max_column,
                FullAnnotationExcelConstants.ATTRIBUTE_STEP,
            ):
                label = ann_sheet.cell(
                    row=row_id, column=col + FullAnnotationExcelConstants.LABEL_OFFSET
                ).value
                if label is None:
                    continue
                value = ann_sheet.cell(
                    row=row_id, column=col + FullAnnotationExcelConstants.VALUE_OFFSET
                ).value
                type = ann_sheet.cell(
                    row=row_id, column=col + FullAnnotationExcelConstants.TYPE_OFFSET
                ).value
                self._fill_attribute(sen, label, value, type, annotator)

    def _fill_modality(self, doc, sen_id, modality, annotator):
        if modality is None:
            logging.debug(f"Modality is left empty for {sen_id}")
            return
        full_annotation = doc[DocumentFields.SENS][sen_id][SenFields.FULL_ANNOTATED_ATTRIBUTES]
        if FullAnnotatedAttributeFields.MODALITY not in full_annotation:
            full_annotation[FullAnnotatedAttributeFields.MODALITY] = {}
        modalities = full_annotation[FullAnnotatedAttributeFields.MODALITY]
        if modality not in modalities:
            modalities[modality] = {AnnotatedAttributeFields.ANNOTATORS: [annotator]}
        else:
            modalities[modality][AnnotatedAttributeFields.ANNOTATORS].append(annotator)

    def _fill_attribute(self, sen, label, value, type, annotator):
        if type not in [
            AttributeTypes.CONDITION,
            AttributeTypes.CONTENT,
            AttributeTypes.CONTENT_EXCEPTION,
            AttributeTypes.CONDITION_EXCEPTION,
        ]:
            logging.warning(f"No such type exists: {type}. Type is set to {EMPTY}")
            type = EMPTY
        annotated_attributes = self.__get_annotated_attributes(sen)
        annotated_values = self.__get_values_for_label(annotated_attributes, label)
        annotated_types = self.__get_types_for_value(annotated_values, value)
        annotators = self.__get_annotators_for_type(annotated_types, type)
        self.__add_annotator(annotators, annotator)

    def __get_annotated_attributes(self, sen):
        full_annotation = sen[SenFields.FULL_ANNOTATED_ATTRIBUTES]
        if FullAnnotatedAttributeFields.ATTRIBUTES not in full_annotation:
            full_annotation[FullAnnotatedAttributeFields.ATTRIBUTES] = {}
        return full_annotation[FullAnnotatedAttributeFields.ATTRIBUTES]

    def __get_values_for_label(self, annotated_attributes, label):
        if label not in annotated_attributes:
            annotated_attributes[label] = {AttributeFields.VALUE: {}}
        return annotated_attributes[label][AttributeFields.VALUE]

    def __get_types_for_value(self, annotated_values, value):
        if value is None:
            value = EMPTY
        if value not in annotated_values:
            annotated_values[value] = {
                AttributeFields.TYPE: {},
            }
        return annotated_values[value][AttributeFields.TYPE]

    def __get_annotators_for_type(self, annotated_types, type):
        if type not in annotated_types:
            annotated_types[type] = {
                AnnotatedAttributeFields.ANNOTATORS: [],
            }
        return annotated_types[type][AnnotatedAttributeFields.ANNOTATORS]

    def __add_annotator(self, annotators, annotator):
        if annotator not in annotators:
            annotators.append(annotator)

    def _fill_with_gold(self, doc):
        for sen in doc[DocumentFields.SENS].values():
            full_attrs_from_gold_sen(sen, self.sen_to_attr, False)

    def _generate_review_excel(self, doc, output_file):
        if not self.review:
            logging.info(
                f"Review = false for {doc[DocumentFields.ID]}, no review excel will be generated."
            )
            return
        generator = FullReviewExcelGenerator(output_file, FullReviewExcelConstants(), doc)
        generator.generate_excel()


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-of", "--output-file", type=str)
    parser.add_argument("-a", "--annotations", nargs="+", default=None)
    parser.add_argument("-d", "--data-file", type=str, default=None)
    parser.add_argument("-g", "--gold-folder", type=str, default=None)
    parser.add_argument("-r", "--review", default=False, action="store_true")
    parser.set_defaults(input_format="XLSX", output_format="XLSX", gen_attributes=False)
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " + "%(module)s (%(lineno)s) - %(levelname)s - %(message)s",
    )
    logging.getLogger("penman").setLevel(logging.WARNING)
    logging.getLogger("stanza").setLevel(logging.WARNING)
    logging.getLogger("tuw_nlp").setLevel(logging.WARNING)
    args = get_args()
    converter = FullAnnotationConverter(args)
    converter.convert(args.annotations, args.output_file, args.data_file)


if __name__ == "__main__":
    main()
