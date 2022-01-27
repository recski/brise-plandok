
import argparse

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from brise_plandok.annotation_process.utils.constants import FullAnnotationExcelConstants
from brise_plandok.annotation_process.utils.full_annotation_pre_filler import FullAnnotationPreFiller
from brise_plandok.attrs_from_gold import SenToAttrMap
from brise_plandok.constants import AttributeFields, SenFields
import logging
import os
from brise_plandok.utils import load_json, normalize_attribute_name
from brise_plandok.xlsx.excel_generator import ExcelGenerator

from openpyxl.worksheet.datavalidation import DataValidation
from brise_plandok.annotation.attributes import ATTR_TO_CAT


LABELS_GOLD = "lables_gold"
FULL_GOLD = "full_gold"


class FullAnnotationExcelGenerator(ExcelGenerator):

    def __init__(self, output_file, CONSTANTS, sen_to_gold_attrs=None):
        self.input_template = os.path.join(os.path.dirname(
            __file__), "../input", "annotation_full_template.xlsx")
        self.output_file = output_file
        self.sen_to_gold_attrs = sen_to_gold_attrs
        self.CONSTANTS = CONSTANTS

    def _modify_header(self, sheet, doc):
        return

    def _fill_modality(self, sen, sheet, row):
        return

    def _gen_attributes(self, sen):
        for attribute in sen[SenFields.GEN_ATTRIBUTES_ON_FULL_ANNOTATION].values():
            if attribute[AttributeFields.VALUE] == []:
                yield {
                    AttributeFields.NAME: attribute[AttributeFields.NAME],
                    AttributeFields.VALUE: None,
                    AttributeFields.TYPE: attribute[AttributeFields.TYPE],
                    LABELS_GOLD: sen[SenFields.LABELS_GOLD_EXISTS],
                    FULL_GOLD: sen[SenFields.FULL_GOLD_EXISTS],
                }
            for value in attribute[AttributeFields.VALUE]:
                yield {
                    AttributeFields.NAME: attribute[AttributeFields.NAME],
                    AttributeFields.VALUE: str(value),
                    AttributeFields.TYPE: attribute[AttributeFields.TYPE],
                    LABELS_GOLD: sen[SenFields.LABELS_GOLD_EXISTS],
                    FULL_GOLD: sen[SenFields.FULL_GOLD_EXISTS],
                }

    def _fill_attribute(self, attribute, sen, sheet, col, row):
        sheet.cell(
            row=row, column=col+self.CONSTANTS.CATEGORY_OFFSET).value = ATTR_TO_CAT[attribute[AttributeFields.NAME]]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.LABEL_OFFSET).value = attribute[AttributeFields.NAME]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.VALUE_OFFSET).value = attribute[AttributeFields.VALUE]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.TYPE_OFFSET).value = attribute[AttributeFields.TYPE]
        self._add_coloring(sheet, col, row, attribute[LABELS_GOLD], attribute[FULL_GOLD])

    def _add_coloring(self, sheet, col, row, lables_gold, full_gold):
        if lables_gold:
            self._color_gold(sheet, row, col)
            self._color_gold(sheet, row, col + self.CONSTANTS.LABEL_OFFSET)
        if full_gold:
            self._color_gold(sheet, row, col + self.CONSTANTS.VALUE_OFFSET)
            self._color_gold(sheet, row, col + self.CONSTANTS.TYPE_OFFSET)

    def _add_validation(self, sheet):
        category_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_NAMED_RANGE}")
        sheet.add_data_validation(category_val)
        type_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.TYPE_NAMED_RANGE}")
        sheet.add_data_validation(type_val)
        modality_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.MODALITY_NAMED_RANGE}")
        sheet.add_data_validation(modality_val)
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, sheet.max_row + 1):
            self._add_validation_for_row(
                sheet, row, category_val, type_val, modality_val)

    def _add_validation_for_row(self, sheet, row, category_val, type_val, modality_val):
        for col in range(1, column_index_from_string(coordinate_from_string(self.CONSTANTS.LAST_COLUMN)[0])):
            if self.__is_modality_cell(col):
                modality_val.add(sheet.cell(row=row, column=col))
            if self._is_category_cell(col):
                self._add_validations_for_attribute(
                    sheet, row, col, category_val)
                type_val.add(sheet.cell(row=row, column=col +
                             self.CONSTANTS.TYPE_OFFSET))

    def __is_modality_cell(self, col):
        return col == self.CONSTANTS.MODALITY_COL


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-o", "--output-file", type=str)
    parser.add_argument("-d", "--data-file", type=str, default=None)
    parser.add_argument("-df", "--data-folder", type=str, default=None)
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " +
               "%(module)s (%(lineno)s) - %(levelname)s - %(message)s")
    args = get_args()
    doc = load_json(args.data_file)
    generator = FullAnnotationExcelGenerator(
        args.output_file, FullAnnotationExcelConstants())
    prefiller = FullAnnotationPreFiller()
    sen_to_gold_attrs = SenToAttrMap(
            gold_dir=args.data_folder, fuzzy=True, full=False) if args.data_folder else None
    prefiller.pre_fill_gold_labels(doc, sen_to_gold_attrs)
    prefiller.fill_gen_attributes_for_full(doc)
    generator.generate_excel(doc)


if __name__ == "__main__":
    main()
