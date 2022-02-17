import argparse
import logging
import os

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.datavalidation import DataValidation

from brise_plandok.annotation.attributes import ATTR_TO_CAT
from brise_plandok.annotation_process.utils.constants import FullReviewExcelConstants, REVIEW_DONE_FLAG
from brise_plandok.constants import EMPTY, AnnotatedAttributeFields, AttributeFields, DocumentFields, \
    FullAnnotatedAttributeFields, SenFields
from brise_plandok.full_attribute_extraction.attribute.potato.potato_predictor import PotatoPredictor
from brise_plandok.full_attribute_extraction.type.extract_types import extract_type
from brise_plandok.full_attribute_extraction.value.extract_values import extract_value
from brise_plandok.utils import load_json
from brise_plandok.xlsx.excel_generator import ExcelGenerator

LABELS_GOLD = "lables_gold"
FULL_GOLD = "full_gold"
PREDICTED_ATTR = "predicted_attr"


class FullReviewExcelGenerator(ExcelGenerator):

    def __init__(self, output_file, constants, doc, sen_to_gold_attrs=None):
        super().__init__(output_file, constants, doc, sen_to_gold_attrs)
        self.input_template = os.path.join(os.path.dirname(
            __file__), "../input", "annotation_full_review_template.xlsx")
        self.potato_predictor = PotatoPredictor(doc)

    def _modify_header(self, sheet):
        self.annotators = self.doc[DocumentFields.FULL_ANNOTATORS]
        if len(self.annotators) > 0:
            sheet.cell(
                row=1, column=self.CONSTANTS.MODALITY_ANN_1_COL).value = self.annotators[0]
        if len(self.annotators) > 1:
            sheet.cell(
                row=1, column=self.CONSTANTS.MODALITY_ANN_2_COL).value = self.annotators[1]
        for col in range(self.CONSTANTS.ATTRIBUTE_OFFSET,
                         column_index_from_string(coordinate_from_string(self.CONSTANTS.LAST_COLUMN)[0]),
                         self.CONSTANTS.ATTRIBUTE_STEP):
            att_count_suffix = " - " + str(self.__get_attribute_count(col))
            sheet.cell(
                row=1, column=col + self.CONSTANTS.CATEGORY_OFFSET).value += att_count_suffix
            sheet.cell(
                row=1, column=col + self.CONSTANTS.LABEL_OFFSET).value += att_count_suffix
            sheet.cell(
                row=1, column=col + self.CONSTANTS.VALUE_OFFSET).value += att_count_suffix
            if len(self.annotators) > 0:
                sheet.cell(
                    row=1, column=col + self.CONSTANTS.TYPE_ANN_1_OFFSET).value = self.annotators[0]
            if len(self.annotators) > 1:
                sheet.cell(
                    row=1, column=col + self.CONSTANTS.TYPE_ANN_2_OFFSET).value = self.annotators[1]

    def __get_attribute_count(self, col):
        return int(((col - self.CONSTANTS.ATTRIBUTE_OFFSET) / self.CONSTANTS.ATTRIBUTE_STEP) + 1)

    def _fill_modality(self, sen, sheet, row):
        ann_modality_exists = FullAnnotatedAttributeFields.MODALITY in sen[SenFields.FULL_ANNOTATED_ATTRIBUTES]
        ann_not_empty = sen[SenFields.FULL_ANNOTATED_ATTRIBUTES] != {}

        if sen[SenFields.FULL_GOLD_EXISTS]:
            sheet.cell(
                row=row, column=self.CONSTANTS.SEN_REVIEW_COL).value = REVIEW_DONE_FLAG
            sheet.cell(
                row=row, column=self.CONSTANTS.MODALITY_ANN_REV_COL).value = sen[SenFields.GOLD_MODALITY]
        elif ann_not_empty and ann_modality_exists:
            ann_modalities = [None] * len(self.annotators)
            for i, annotator in enumerate(self.annotators):
                raw_modalities = sen[SenFields.FULL_ANNOTATED_ATTRIBUTES][FullAnnotatedAttributeFields.MODALITY].items()
                for modality, annotators in raw_modalities:
                    if annotator in annotators[AnnotatedAttributeFields.ANNOTATORS]:
                        ann_modalities[i] = modality
            sheet.cell(
                row=row, column=self.CONSTANTS.MODALITY_ANN_1_COL).value = ann_modalities[0]
            sheet.cell(
                row=row, column=self.CONSTANTS.MODALITY_ANN_2_COL).value = ann_modalities[1]
            if len(ann_modalities) > 1 and ann_modalities[0] == ann_modalities[1]:
                sheet.cell(
                    row=row, column=self.CONSTANTS.MODALITY_ANN_REV_COL).value = ann_modalities[0]

    def _gen_attributes(self, sen):
        if sen[SenFields.FULL_GOLD_EXISTS]:
            yield from self.__gen_attributes_from_gold(sen)
        else:
            ann_attributes = {}
            if FullAnnotatedAttributeFields.ATTRIBUTES in sen[SenFields.FULL_ANNOTATED_ATTRIBUTES]:
                ann_attributes = sen[SenFields.FULL_ANNOTATED_ATTRIBUTES][FullAnnotatedAttributeFields.ATTRIBUTES]
            potato_predictions = self.potato_predictor.get_prediction_for_sen(sen)
            yield from self.__gen_attributes_from_annotation(sen, ann_attributes, potato_predictions)
            yield from self.__add_attributes_from_prediction(sen, ann_attributes, potato_predictions)

    def __gen_attributes_from_gold(self, sen):
        for full_attribute in sen[SenFields.GOLD_ATTRIBUTES].values():
            ann_types = [None] * (len(self.annotators) + 1)
            name = full_attribute[AttributeFields.NAME]
            values = full_attribute[AttributeFields.VALUE]
            types = full_attribute[AttributeFields.TYPE]
            if len(values) != len(types):
                raise ValueError(
                    f"{sen[SenFields.ID]}: Lenght of values {len(values)} is not equal to lenght of types {len(types)}")
            for i in range(len(values)):
                ann_types[-1] = types[i]
                yield {
                    AttributeFields.NAME: name,
                    AttributeFields.VALUE: values[i],
                    AttributeFields.TYPE: ann_types,
                    LABELS_GOLD: True,
                    FULL_GOLD: True,
                    PREDICTED_ATTR: False,
                }

    def __gen_attributes_from_annotation(self, sen, ann_attributes, attribute_predictions):
        for attribute_name, attribute in ann_attributes.items():
            ann_types = [None] * len(self.annotators)
            for value_name, annotated_types in attribute[AttributeFields.VALUE].items():
                for i, annotator in enumerate(self.annotators):
                    try:
                        ann_types[i] = self.__get_type_for_annotator(sen[SenFields.ID],
                                                                     annotated_types[AttributeFields.TYPE],
                                                                     annotator)
                    except ValueError as err:
                        logging.error(err)
                        logging.error(f"Conflict in {sen[SenFields.FULL_ANNOTATED_ATTRIBUTES]}")
                yield {
                    AttributeFields.NAME: attribute_name,
                    AttributeFields.VALUE: "\n".join(
                        extract_value(sen, attribute_name, field_to_add=None, only_gold=False)),
                    AttributeFields.TYPE: ann_types,
                    LABELS_GOLD: sen[SenFields.LABELS_GOLD_EXISTS],
                    FULL_GOLD: sen[SenFields.FULL_GOLD_EXISTS],
                    PREDICTED_ATTR: attribute_name in attribute_predictions
                }

    def __add_attributes_from_prediction(self, sen, ann_attributes, attribute_predictions):
        for attribute_name in attribute_predictions:
            if attribute_name not in ann_attributes.keys():
                extracted_type = extract_type(sen, attribute_name, field_to_add=None, only_gold=False)
                yield {
                    AttributeFields.NAME: attribute_name,
                    AttributeFields.VALUE: "\n".join(
                        extract_value(sen, attribute_name, field_to_add=None, only_gold=False)),
                    AttributeFields.TYPE: [extracted_type] * (len(self.annotators)),
                    LABELS_GOLD: False,
                    FULL_GOLD: False,
                    PREDICTED_ATTR: True,
                }

    def __get_type_for_annotator(self, sen_id, annotated_types, annotator):
        types_from_ann = []
        for ann_type, annotators in annotated_types.items():
            if annotator in annotators[AnnotatedAttributeFields.ANNOTATORS]:
                types_from_ann.append(ann_type)
        if len(types_from_ann) > 1:
            raise ValueError(
                f"Annotator {annotator} gave different types for same label with same value for {sen_id}: {types_from_ann}")
        if len(types_from_ann) == 1:
            return types_from_ann[0]
        return None

    def _fill_attribute(self, attribute, sen, sheet, col, row):
        sheet.cell(
            row=row, column=col + self.CONSTANTS.CATEGORY_OFFSET).value = ATTR_TO_CAT[
            attribute[AttributeFields.NAME]]
        sheet.cell(row=row, column=col + self.CONSTANTS.LABEL_OFFSET).value = attribute[AttributeFields.NAME]
        sheet.cell(row=row, column=col + self.CONSTANTS.VALUE_OFFSET).value = attribute[AttributeFields.VALUE]
        types = attribute[AttributeFields.TYPE]
        if attribute[FULL_GOLD]:
            sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_REV_OFFSET).value = types[-1]
        elif types is not None:
            if len(types) > 0:
                sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_1_OFFSET).value = types[0]
            if len(types) > 1:
                sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_2_OFFSET).value = types[1]
                if types[0] == types[1] and types[0] != EMPTY:
                    sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_REV_OFFSET).value = types[0]
        self._add_coloring(
            sheet, col, row, attribute[LABELS_GOLD], attribute[FULL_GOLD], attribute[PREDICTED_ATTR])

    def _add_coloring(self, sheet, col, row, labels_gold, full_gold, predicted):
        if labels_gold or full_gold:
            if labels_gold:
                self._color_gold(sheet, row, col)
                self._color_gold(sheet, row, col + self.CONSTANTS.LABEL_OFFSET)
            if full_gold:
                self._color_gold(sheet, row, col + self.CONSTANTS.VALUE_OFFSET)
        elif predicted:
            self._color_gray(sheet, row, col)
            self._color_gray(sheet, row, col + self.CONSTANTS.LABEL_OFFSET)
            self._color_gray(sheet, row, col + self.CONSTANTS.VALUE_OFFSET)

    def _add_validation(self, sheet):
        category_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_NAMED_RANGE}")
        sheet.add_data_validation(category_val)
        type_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.TYPE_NAMED_RANGE}")
        sheet.add_data_validation(type_val)
        modality_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.MODALITY_NAMED_RANGE}")
        sheet.add_data_validation(modality_val)
        sen_review_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.SENTENCE_REVIEW_NAMED_RANGE}")
        sheet.add_data_validation(sen_review_val)
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, sheet.max_row + 1):
            self._add_validation_for_row(
                sheet, row, category_val, type_val, modality_val, sen_review_val)

    def _add_validation_for_row(self, sheet, row, category_val, type_val, modality_val, sen_review_val):
        sen_review_val.add(sheet.cell(row=row, column=self.CONSTANTS.SEN_REVIEW_COL))
        modality_val.add(sheet.cell(row=row, column=self.CONSTANTS.MODALITY_ANN_1_COL))
        modality_val.add(sheet.cell(row=row, column=self.CONSTANTS.MODALITY_ANN_2_COL))
        modality_val.add(sheet.cell(row=row, column=self.CONSTANTS.MODALITY_ANN_REV_COL))
        for col in range(self.CONSTANTS.ATTRIBUTE_OFFSET,
                         column_index_from_string(coordinate_from_string(self.CONSTANTS.LAST_COLUMN)[0])):
            if self._is_category_cell(col):
                self._add_validations_for_attribute(
                    sheet, row, col, category_val)
                type_val.add(sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_1_OFFSET))
                type_val.add(sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_2_OFFSET))
                type_val.add(sheet.cell(row=row, column=col + self.CONSTANTS.TYPE_ANN_REV_OFFSET))


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-o", "--output-file", type=str)
    parser.add_argument("-d", "--data-file", type=str, default=None)
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " +
               "%(module)s (%(lineno)s) - %(levelname)s - %(message)s")
    args = get_args()
    doc = load_json(args.data_file)
    generator = FullReviewExcelGenerator(
        args.output_file, FullReviewExcelConstants(), doc)
    generator.generate_excel()


if __name__ == "__main__":
    main()
