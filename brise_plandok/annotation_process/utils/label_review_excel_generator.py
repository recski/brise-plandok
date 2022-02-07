from brise_plandok.constants import AttributeFields, Review, SenFields
from brise_plandok.constants import SenFields as SF
from brise_plandok.constants import AnnotatedAttributeFields as AAF
import logging
import os
from brise_plandok.data_utils import normalize_attribute_name
from brise_plandok.xlsx.excel_generator import ExcelGenerator

from openpyxl.worksheet.datavalidation import DataValidation
from brise_plandok.annotation.attributes import ATTR_TO_CAT

IS_GOLD = "gold_attr"


class LabelReviewExcelGenerator(ExcelGenerator):

    def __init__(self, output_file, CONSTANTS, sen_to_gold_attrs=None):
        self.input_template = os.path.join(os.path.dirname(
            __file__), "../input", "labels_review_template.xlsx")
        self.output_file = output_file
        self.sen_to_gold_attrs = sen_to_gold_attrs
        self.CONSTANTS = CONSTANTS

    def _modify_header(self, sheet, doc):
        return

    def _fill_modality(self, sen, sheet, row):
        return

    def _gen_attributes(self, sen):
        return [{AttributeFields.NAME: attr} for attr in list(sen[SF.ANNOTATED_ATTRIBUTES].keys()) + list(sen[SF.GOLD_ATTRIBUTES].keys())]

    def _fill_attribute(self, attribute, sen, sheet, col, row):
        attribute_name = attribute[AttributeFields.NAME]
        sheet.cell(
            row=row, column=col+self.CONSTANTS.CATEGORY_OFFSET).value = ATTR_TO_CAT[attribute_name]
        sheet.cell(row=row, column=col +
                          self.CONSTANTS.LABEL_OFFSET).value = attribute_name
        self.__set_annotator_count(attribute_name, sen, sheet, row, col)
        self.__set_annotator_text(attribute_name, sen, sheet, row, col)
        self.__set_review_value(attribute_name, sen, sheet, row, col)

    def __set_annotator_count(self, attribute_name, sen, sheet, row, col):
        count = 0
        if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
            logging.info(
                f"{sen[SenFields.ID]}: no annotator found - setting count for {attribute_name} to 0")
        else:
            count = len(sen[SF.ANNOTATED_ATTRIBUTES]
                        [attribute_name][AAF.ANNOTATORS])
        sheet.cell(row=row, column=col +
                          self.CONSTANTS.COUNT_OFFSET).value = count

    def __set_annotator_text(self, attribute_name, sen, sheet, row, col):
        annotators = ""
        if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
            logging.info(
                f"{sen[SenFields.ID]}: no annotator found - setting annotator for {attribute_name} to gold")
            annotators = "gold"
        else:
            annotators = self.CONSTANTS.ANNOTATOR_SEPARATOR.join(
                sen[SF.ANNOTATED_ATTRIBUTES][attribute_name][AAF.ANNOTATORS])
        sheet.cell(row=row, column=col +
                          self.CONSTANTS.ANNOTATORS_OFFSET).value = annotators

    def __set_review_value(self, attribute_name, sen, sheet, row, col):
        review_value = Review.OK
        if self._labels_gold_exists(sen):
            review_value = self.__set_gold_review_value(
                attribute_name, sen, sheet, row, col)
        elif attribute_name in sen[SF.GEN_ATTRIBUTES_ON_ANNOTATION]:
            self._color_gray(sheet, row, col +
                             self.CONSTANTS.CATEGORY_OFFSET)
            self._color_gray(sheet, row, col +
                             self.CONSTANTS.LABEL_OFFSET)
            self._color_gray(sheet, row, col +
                             self.CONSTANTS.COUNT_OFFSET)
            self._color_gray(sheet, row, col +
                             self.CONSTANTS.ANNOTATORS_OFFSET)
        sheet.cell(row=row, column=col +
                          self.CONSTANTS.ATTRIBUTE_REVIEW_OFFSET).value = review_value

    def __set_gold_review_value(self, attribute_name, sen, sheet, row, col):
        if attribute_name not in sen[SF.GOLD_ATTRIBUTES]:
            return Review.ERROR
        else:
            self._color_gold(sheet, row, col +
                             self.CONSTANTS.CATEGORY_OFFSET)
            self._color_gold(sheet, row, col +
                             self.CONSTANTS.LABEL_OFFSET)
            self._color_gold(sheet, row, col +
                             self.CONSTANTS.COUNT_OFFSET)
            self._color_gold(sheet, row, col +
                             self.CONSTANTS.ANNOTATORS_OFFSET)
            if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
                return Review.MISSING
        return Review.OK

    def _add_validation(self, sheet):
        data_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_NAMED_RANGE}")
        sheet.add_data_validation(data_val)
        review_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_REVIEW_NAMED_RANGE}")
        sheet.add_data_validation(review_val)
        sen_review_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.SENTENCE_REVIEW_NAMED_RANGE}")
        sheet.add_data_validation(sen_review_val)
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, sheet.max_row + 1):
            self.__add_validation_for_row(
                sheet, row, data_val, review_val, sen_review_val)

    def __add_validation_for_row(self, sheet, row, data_val, review_val, sen_review_val):
        for cell in sheet[str(row)]:
            col = cell.column
            if self.__is_sentence_review_cell(col):
                self.__add_validation_for_sentence_review(
                    sheet, row, col, sen_review_val)
            elif self._is_category_cell(col):
                self._add_validations_for_attribute(
                    sheet, row, col, data_val)
                review_val.add(
                    sheet.cell(row=row, column=col+self.CONSTANTS.ATTRIBUTE_REVIEW_OFFSET))

    def __add_validation_for_sentence_review(self, sheet, row, col, sentence_review_val):
        sentence_review_val.add(sheet.cell(row=row, column=col))

    def __is_sentence_review_cell(self, col):
        return col == self.CONSTANTS.SEN_REVIEW_COL
