import json
import logging

import openpyxl

from brise_plandok.constants import DocumentFields, SenFields
from brise_plandok.utils import dump_json, load_json, update_gold_docs


class ReviewConverter:
    def __init__(self, data_file, gold_folder):
        self.data = load_json(data_file)
        self.data_file = data_file
        self.gold_folder = gold_folder
        self.SEN_GOLD = None
        self.DOC_GOLD = None
        self.CONSTANTS = None
        self.sen_to_gold_attrs = None

    def convert(self, reviewed_xlsx_file, overwrite_internal, overwrite_external):
        self._fill_gold_attrs(reviewed_xlsx_file, overwrite_internal, overwrite_external)
        self._save_gold_json()

    def _save_gold_json(self):
        dump_json(self.data, self.data_file)
        logging.info(f"DONE. Gold json was created to: {self.data_file}")

    def _fill_gold_attrs(self, fn, overwrite_internal, overwrite_external):
        workbook = openpyxl.load_workbook(fn)
        review_sheet = workbook[self.CONSTANTS.MAIN_SHEET_NAME]

        if overwrite_internal:
            logging.info("Internal overwrite is true, no internal checks will be done.")
        if overwrite_external:
            logging.info("External overwrite is true, other docs will be overwritten.")

        for row_id in range(self.CONSTANTS.FIRST_DATA_ROW, review_sheet.max_row + 1):
            sen_id = review_sheet.cell(row=row_id, column=self.CONSTANTS.SEN_ID_COL).value

            attributes = [
                attribute for attribute in self._generate_attributes(review_sheet, row_id)
            ]
            gold_modality_candidate = self._get_modality(review_sheet, row_id, attributes)

            if self._is_error(attributes):
                self.data[DocumentFields.SENS][sen_id][SenFields.SEGMENTATION_ERROR] = True
                logging.info(f"error row found '{sen_id}' - skipping from gold")
                continue

            gold_attr_candidate = self._get_gold_candidate(sen_id, attributes)
            self._handle_internal_conflict(
                sen_id, gold_attr_candidate, gold_modality_candidate, overwrite_internal
            )
            self._handle_external_conflict(
                sen_id, gold_attr_candidate, gold_modality_candidate, overwrite_external
            )

            self.data[DocumentFields.SENS][sen_id][self.SEN_GOLD] = True
            self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_ATTRIBUTES] = gold_attr_candidate
            self.data[DocumentFields.SENS][sen_id][
                SenFields.GOLD_MODALITY
            ] = gold_modality_candidate

        self.data[self.DOC_GOLD] = True

    def _generate_attributes(self, review_sheet, row_id):
        raise NotImplementedError()

    def _get_gold_candidate(self, sen_id, attributes):
        raise NotImplementedError()

    def _get_modality(self, review_sheet, row_id, attributes):
        raise NotImplementedError()

    def _handle_internal_conflict(
        self, sen_id, gold_attr_candidate, gold_mod_candidate, overwrite
    ):
        if not overwrite:
            if self.data[DocumentFields.SENS][sen_id][self.SEN_GOLD]:
                current_gold_attr = self.data[DocumentFields.SENS][sen_id][
                    SenFields.GOLD_ATTRIBUTES
                ]
                current_gold_mod = self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_MODALITY]
                if gold_attr_candidate != current_gold_attr:
                    logging.error(
                        f"Conflict within already gold sentence {sen_id}:\n"
                        f"Current ({sen_id}):\n"
                        f"{json.dumps(current_gold_attr, indent=2)}\n"
                        f"New:\n"
                        f"{json.dumps(gold_attr_candidate, indent=2)}"
                    )
                    raise ValueError("Gold conflict with attributes")
                if gold_mod_candidate != current_gold_mod:
                    logging.error(
                        f"Conflict within already gold sentence {sen_id}:\n"
                        f"Current ({sen_id}):\n"
                        f"{json.dumps(current_gold_mod, indent=2)}\n"
                        f"New:\n"
                        f"{json.dumps(gold_mod_candidate, indent=2)}"
                    )
                    raise ValueError("Gold conflict with modality")

    def _handle_external_conflict(
        self, sen_id, gold_attr_candidate, gold_mod_candidate, overwrite
    ):
        text = self.data[DocumentFields.SENS][sen_id][SenFields.TEXT]
        current_gold_sens = self.sen_to_gold_attrs.get_sens(text)
        if current_gold_sens != [sen_id]:
            current_gold_attr = self.sen_to_gold_attrs.get_attrs(text)
            current_gold_mod = self.sen_to_gold_attrs.get_mod(text)
            update_docs = False
            update_docs = self._handle_external_attribute_conflict(
                current_gold_attr,
                current_gold_sens,
                gold_attr_candidate,
                overwrite,
                sen_id,
                text,
                update_docs,
            )
            update_docs = self._handle_external_modality_conflict(
                current_gold_mod,
                current_gold_sens,
                gold_mod_candidate,
                overwrite,
                sen_id,
                text,
                update_docs,
            )
            if update_docs:
                update_gold_docs(gold_attr_candidate, gold_mod_candidate, current_gold_sens)

    def _handle_external_attribute_conflict(
        self,
        current_gold_attr,
        current_gold_sens,
        gold_attr_candidate,
        overwrite,
        sen_id,
        text,
        update_docs,
    ):
        if current_gold_attr is not None:
            if gold_attr_candidate != current_gold_attr:
                logging.info(
                    f"Conflict with current gold value of {sen_id}:\nCurrent ({current_gold_sens}):\n{current_gold_attr}\nNew:\n{gold_attr_candidate}"
                )
                if overwrite:
                    self.sen_to_gold_attrs.set_attrs(text, gold_attr_candidate)
                    update_docs = True
                else:
                    raise ValueError("Gold conflict with attributes")
        return update_docs

    def _handle_external_modality_conflict(
        self,
        current_gold_mod,
        current_gold_sens,
        gold_mod_candidate,
        overwrite,
        sen_id,
        text,
        update_docs,
    ):
        if current_gold_mod is not None:
            if gold_mod_candidate != current_gold_mod:
                logging.info(
                    f"Conflict with current gold value of {sen_id}:\nCurrent ({current_gold_sens}):\n{current_gold_mod}\nNew:\n{gold_mod_candidate}"
                )
                if overwrite:
                    self.sen_to_gold_attrs.set_mod(text, gold_mod_candidate)
                    update_docs = True
                else:
                    raise ValueError("Gold conflict with modality")
        return update_docs

    def _is_error(self, attributes):
        return self.CONSTANTS.ERROR_LABEL in set([attr["name"] for attr in attributes])
