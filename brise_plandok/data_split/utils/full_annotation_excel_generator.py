
import argparse

from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from brise_plandok.constants import AttributeFields, SenFields
import logging
import os
from brise_plandok.data_split.utils.constants import FullAnnotationExcelConstants
from brise_plandok.utils import load_json, normalize_attribute_name
from brise_plandok.xlsx.excel_generator import ExcelGenerator

from openpyxl.worksheet.datavalidation import DataValidation
from brise_plandok.annotation.attributes import ATTR_TO_CAT


IS_GOLD = "gold_attr"


class FullAnnotationExcelGenerator(ExcelGenerator):

    def __init__(self, output_file, CONSTANTS, sen_to_gold_attrs=None):
        self.input_template = os.path.join(os.path.dirname(
            __file__), "../input", "annotation_phase2_template.xlsx")
        self.output_file = output_file
        self.sen_to_gold_attrs = sen_to_gold_attrs
        self.CONSTANTS = CONSTANTS

    def _fill_attributes(self, sen, sheet, row):
        col = self.CONSTANTS.ATTRIBUTE_OFFSET
        for attribute in self.__gen_one_attribute_per_value(sen):
            attribute_name = normalize_attribute_name(
                attribute[AttributeFields.NAME])
            if attribute_name not in ATTR_TO_CAT:
                logging.warn(
                    f"\"{attribute_name}\" does not belong to any category - will be ignored")
            else:
                self._fill_attribute(
                    attribute, sen, sheet, col, row)
                col += self.CONSTANTS.ATTRIBUTE_STEP

    def __gen_one_attribute_per_value(self, sen):
        for attribute in sen[SenFields.GEN_ATTRIBUTES_ON_FULL_ANNOTATION].values():
            if attribute[AttributeFields.VALUE] == []:
                yield attribute
            for value in attribute[AttributeFields.VALUE]:
                yield {
                    AttributeFields.NAME: attribute[AttributeFields.NAME],
                    AttributeFields.VALUE: value,
                    AttributeFields.TYPE: attribute[AttributeFields.TYPE],
                }

    def _fill_attribute(self, attribute, sen, sheet, col, row):
        sheet.cell(
            row=row, column=col+self.CONSTANTS.CATEGORY_OFFSET).value = ATTR_TO_CAT[attribute[AttributeFields.NAME]]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.LABEL_OFFSET).value = attribute[AttributeFields.NAME]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.VALUE_OFFSET).value = attribute[AttributeFields.VALUE]
        sheet.cell(row=row, column=col +
                   self.CONSTANTS.TYPE_OFFSET).value = attribute[AttributeFields.TYPE]
        self._fill_value(attribute, sen, sheet, row, col)
        self._fill_type(attribute, sen, sheet, row, col)

    def _add_validation(self, sheet):
        category_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_NAMED_RANGE}")
        sheet.add_data_validation(category_val)
        type_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.TYPE_NAMED_RANGE}")
        sheet.add_data_validation(type_val)
        modality_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.MODALITY_NAMED_RANGE}")
        sheet.add_data_validation(modality_val)
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, sheet.max_row + 1):
            self._add_validation_for_row(
                sheet, row, category_val, type_val, modality_val)

    def _add_validation_for_row(self, sheet, row, category_val, type_val, modality_val):
        for col in range(1, column_index_from_string(coordinate_from_string(self.CONSTANTS.LAST_COLUMN)[0])):
            if self.__is_modality_cell(col):
                modality_val.add(sheet.cell(row=row, column=col))
            if self._is_category_cell(col):
                self._add_validations_for_attribute(
                    sheet, row, col, category_val)
                type_val.add(sheet.cell(row=row, column=col+self.CONSTANTS.TYPE_OFFSET))

    def __is_modality_cell(self, col):
        return col == self.CONSTANTS.MODALITY_COL


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-o", "--output-file", type=str)
    parser.add_argument("-d", "--data-file", type=str, default=None)
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " +
               "%(module)s (%(lineno)s) - %(levelname)s - %(message)s")
    args = get_args()
    doc = load_json(args.data_file)
    generator = FullAnnotationExcelGenerator(
        args.output_file, FullAnnotationExcelConstants())
    generator.generate_excel(doc)


if __name__ == "__main__":
    main()
