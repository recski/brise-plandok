import argparse
import json
import os
from brise_plandok.review.constants import ANNOTATORS_OFFSET, ANNOTATOR_SEPARATOR, ATTRIBUTES_TO_IGNORE, ATTRIBUTE_OFFSET, ATTRIBUTE_REVIEW_OFFSET, ATTRIBUTE_STEP, ERROR_LABEL, FIRST_DATA_ROW, LABEL_OFFSET, REVIEW_SHEET_NAME, SEN_ID_COL, SEN_TEXT_COL

import openpyxl
import logging
from brise_plandok.convert import Converter


class ReviewConverter(Converter):

    def __init__(self, output_file):
        self.output_file = output_file

    def convert(self, reviewed_xlsx_file):
        doc = self.__read_reviewed_xlsx(reviewed_xlsx_file)
        self.__save_gold_json(doc)

    def __save_gold_json(self, doc):
        with open(self.output_file, "w") as f:
            json.dump(doc, f)
        logging.info(f"DONE. Gold json was created to: {self.output_file}")

    def __read_reviewed_xlsx(self, reviewed_xlsx_file):
        doc = {
            "id": os.path.basename(reviewed_xlsx_file).split('_')[0],
            "sens": [sen for sen in self.__gen_sens_from_file(reviewed_xlsx_file)],
        }

        return doc

    def __gen_sens_from_file(self, fn):
        workbook = openpyxl.load_workbook(fn)
        review_sheet = workbook[REVIEW_SHEET_NAME]

        for row_id in range(FIRST_DATA_ROW, review_sheet.max_row):
            attributes = [attribute for attribute in self.__generate_attributes(review_sheet, row_id)]
            if not self.__is_error_row(attributes):
                yield {
                    "id": review_sheet.cell(row=row_id, column=SEN_ID_COL).value,
                    "text": review_sheet.cell(row=row_id, column=SEN_TEXT_COL).value,
                    "modality": None,
                    "gold_attributes": [{"name": attribute["name"], "type": None, "value": None} for attribute in attributes if attribute["gold"]],
                    "annotated_attributes": [{"name": attribute["name"], "annotators": attribute["annotators"]} for attribute in attributes],
                }

    def __generate_attributes(self, review_sheet, row_id):
        for col in range(ATTRIBUTE_OFFSET, review_sheet.max_column, ATTRIBUTE_STEP):
            label = review_sheet.cell(
                row=row_id, column=col+LABEL_OFFSET).value
            if label is None:
                continue
            review = review_sheet.cell(
                row=row_id, column=col+ATTRIBUTE_REVIEW_OFFSET).value
            if review is None:
                raise ValueError("Review field is not filled out for row " + str(row_id))
            if label not in ATTRIBUTES_TO_IGNORE:
                yield {
                    "name": label,
                    "annotators": self.__get_annotators(review_sheet, row_id, col),
                    "gold": self.__is_gold(review),
                }

    def __get_annotators(self, review_sheet, row_id, col):
        annotators = review_sheet.cell(
            row=row_id, column=col+ANNOTATORS_OFFSET).value
        if not annotators:
            return None
        return annotators.split(ANNOTATOR_SEPARATOR)

    def __is_gold(self, review):
        return review == "OK" or review == "MISSING"

    def __is_error_row(self, attributes):
        return ERROR_LABEL in set([attr["name"] for attr in attributes])


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-o", "--output-file", type=str)
    parser.add_argument("-r", "--review", default=None)
    parser.set_defaults(input_format="XLSX", output_format="JSON",
                        output_file="brise_plandok/review/output/gold.json")
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " +
               "%(module)s (%(lineno)s) - %(levelname)s - %(message)s")
    args = get_args()
    converter = ReviewConverter(args.output_file)
    converter.convert(args.review)


if __name__ == "__main__":
    main()
