from brise_plandok.constants import Review, SenFields
from brise_plandok.constants import SenFields as SF
from brise_plandok.constants import AnnotatedAttributeFields as AAF
import logging
import os
from brise_plandok.utils import normalize_attribute_name
from brise_plandok.xlsx.excel_generator import ExcelGenerator

from openpyxl.worksheet.datavalidation import DataValidation
from brise_plandok.annotation.attributes import ATTR_TO_CAT

IS_GOLD = "gold_attr"


class ReviewExcelGenerator(ExcelGenerator):

    def __init__(self, output_file, CONSTANTS, sen_to_gold_attrs=None):
        self.input_template = os.path.join(os.path.dirname(
            __file__), "input", "review_template.xlsx")
        self.output_file = output_file
        self.sen_to_gold_attrs = sen_to_gold_attrs
        self.CONSTANTS = CONSTANTS

    def _fill_attributes(self, sen, review_sheet, row):
        col = self.CONSTANTS.ATTRIBUTE_OFFSET
        for attribute_name in self._get_annotated_and_gold_attrs(sen):
            attribute_name = normalize_attribute_name(attribute_name)
            if attribute_name not in ATTR_TO_CAT:
                logging.warn(
                    f"\"{attribute_name}\" does not belong to any category - will be ignored")
            else:
                self._fill_attribute(
                    attribute_name, sen, review_sheet, col, row)
                col += self.CONSTANTS.ATTRIBUTE_STEP

    def _get_annotated_and_gold_attrs(self, sen):
        return set(list(sen[SF.ANNOTATED_ATTRIBUTES].keys()) + list(sen[SF.GOLD_ATTRIBUTES].keys()))

    def _fill_attribute(self, attribute_name, sen, review_sheet, col, row):
        review_sheet.cell(
            row=row, column=col+self.CONSTANTS.CATEGORY_OFFSET).value = ATTR_TO_CAT[attribute_name]
        review_sheet.cell(row=row, column=col +
                          self.CONSTANTS.LABEL_OFFSET).value = attribute_name
        self._set_annotator_count(attribute_name, sen, review_sheet, row, col)
        self._set_annotator_text(attribute_name, sen, review_sheet, row, col)
        self._set_review_value(attribute_name, sen, review_sheet, row, col)

    def _set_annotator_count(self, attribute_name, sen, review_sheet, row, col):
        count = 0
        if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
            logging.info(
                f"{sen[SenFields.ID]}: no annotator found - setting count for {attribute_name} to 0")
        else:
            count = len(sen[SF.ANNOTATED_ATTRIBUTES]
                        [attribute_name][AAF.ANNOTATORS])
        review_sheet.cell(row=row, column=col +
                          self.CONSTANTS.COUNT_OFFSET).value = count

    def _set_annotator_text(self, attribute_name, sen, review_sheet, row, col):
        annotators = ""
        if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
            logging.info(
                f"{sen[SenFields.ID]}: no annotator found - setting annotator for {attribute_name} to gold")
            annotators = "gold"
        else:
            annotators = self.CONSTANTS.ANNOTATOR_SEPARATOR.join(
                sen[SF.ANNOTATED_ATTRIBUTES][attribute_name][AAF.ANNOTATORS])
        review_sheet.cell(row=row, column=col +
                          self.CONSTANTS.ANNOTATORS_OFFSET).value = annotators

    def _set_review_value(self, attribute_name, sen, review_sheet, row, col):
        review_value = Review.OK
        if self._is_gold(sen):
            review_value = self._set_gold_review_value(
                attribute_name, sen, review_sheet, row, col)
        elif attribute_name in sen[SF.GEN_ATTRIBUTES_ON_ANNOTATION]:
            self._color_gray(review_sheet, row, col +
                             self.CONSTANTS.CATEGORY_OFFSET)
            self._color_gray(review_sheet, row, col +
                             self.CONSTANTS.LABEL_OFFSET)
            self._color_gray(review_sheet, row, col +
                             self.CONSTANTS.COUNT_OFFSET)
            self._color_gray(review_sheet, row, col +
                             self.CONSTANTS.ANNOTATORS_OFFSET)
        review_sheet.cell(row=row, column=col +
                          self.CONSTANTS.ATTRIBUTE_REVIEW_OFFSET).value = review_value

    def _set_gold_review_value(self, attribute_name, sen, review_sheet, row, col):
        if attribute_name not in sen[SF.GOLD_ATTRIBUTES]:
            return Review.ERROR
        else:
            self._color_gold(review_sheet, row, col +
                             self.CONSTANTS.CATEGORY_OFFSET)
            self._color_gold(review_sheet, row, col +
                             self.CONSTANTS.LABEL_OFFSET)
            self._color_gold(review_sheet, row, col +
                             self.CONSTANTS.COUNT_OFFSET)
            self._color_gold(review_sheet, row, col +
                             self.CONSTANTS.ANNOTATORS_OFFSET)
            if attribute_name not in sen[SF.ANNOTATED_ATTRIBUTES]:
                return Review.MISSING
        return Review.OK

    def _substitute_with_gold(self, gold_attributes, annotation):
        for attribute_name, attribute_props in annotation["attributes"].items():
            if attribute_name in gold_attributes:
                attribute_props[IS_GOLD] = True
            else:
                attribute_props[IS_GOLD] = False
        for attribute in gold_attributes:
            if attribute not in annotation["attributes"]:
                annotation["attributes"][attribute] = {
                    "count": 0,
                    IS_GOLD: True,
                    "annotators": ["gold"],
                }

    def _add_validation(self, review_sheet):
        data_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_NAMED_RANGE}")
        review_sheet.add_data_validation(data_val)
        review_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.ATTRIBUTE_REVIEW_NAMED_RANGE}")
        review_sheet.add_data_validation(review_val)
        sen_review_val = DataValidation(
            type="list", formula1=f"={self.CONSTANTS.SENTENCE_REVIEW_NAMED_RANGE}")
        review_sheet.add_data_validation(sen_review_val)
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, review_sheet.max_row + 1):
            self._add_validation_for_row(
                review_sheet, row, data_val, review_val, sen_review_val)

    def _add_validation_for_row(self, review_sheet, row, data_val, review_val, sen_review_val):
        for cell in review_sheet[str(row)]:
            col = cell.column
            if self._is_sentence_review_cell(col):
                self._add_validation_for_sentence_review(
                    review_sheet, row, col, sen_review_val)
            elif self._is_category_cell(col):
                self._add_validations_for_attribute(
                    review_sheet, row, col, data_val, review_val)

    def _add_validation_for_sentence_review(self, review_sheet, row, col, sentence_review_val):
        sentence_review_val.add(review_sheet.cell(row=row, column=col))

    def _is_sentence_review_cell(self, col):
        return col == self.CONSTANTS.SEN_REVIEW_COL
