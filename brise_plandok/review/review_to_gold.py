import argparse
from brise_plandok.attrs_from_gold import SenToAttrMap
from brise_plandok.constants import AttributeFields, DocumentFields, SenFields
from utils import dump_json, load_json
from brise_plandok.review.constants import ATTRIBUTES_TO_IGNORE, ATTRIBUTE_OFFSET, ATTRIBUTE_REVIEW_OFFSET, ATTRIBUTE_STEP, ERROR_LABEL, FIRST_DATA_ROW, LABEL_OFFSET, REVIEW_SHEET_NAME, SEN_ID_COL

import openpyxl
import logging
from brise_plandok.convert import Converter


class ReviewConverter(Converter):

    def __init__(self, data_file, gold_folder):
        self.data = load_json(data_file)
        self.data_file = data_file
        self.sen_to_gold_attrs = SenToAttrMap(gold_dir=gold_folder, fuzzy=True)

    def convert(self, reviewed_xlsx_file):
        self._fill_gold_attrs(reviewed_xlsx_file)
        self._save_gold_json()

    def _save_gold_json(self):
        dump_json(self.data, self.data_file)
        logging.info(f"DONE. Gold json was created to: {self.data_file}")

    def _fill_gold_attrs(self, fn):
        workbook = openpyxl.load_workbook(fn)
        review_sheet = workbook[REVIEW_SHEET_NAME]

        for row_id in range(FIRST_DATA_ROW, review_sheet.max_row):
            sen_id = review_sheet.cell(row=row_id, column=SEN_ID_COL).value
            attributes = [attribute for attribute in self._generate_attributes(
                review_sheet, row_id)]
            if self._is_error(attributes):
                self.data[DocumentFields.SENS][sen_id][SenFields.SEGMENTATION_ERROR] = True
                logging.info(
                    f"error row found '{sen_id}' - skipping from gold")
                continue

            gold_candidate = self._get_gold_candidate(attributes)
            self._raise_error_on_internal_conflict(sen_id, gold_candidate)
            self._raise_error_on_external_conflict(sen_id, gold_candidate)

            self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_EXISTS] = True
            self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_ATTRIBUTES] = gold_candidate

        self.data[DocumentFields.IS_GOLD] = True

    def _get_gold_candidate(self, attributes):
        gold_candidate = {}
        for attr_name, attr in self._generate_gold_attrs(attributes):
            assert attr_name not in gold_candidate
            gold_candidate[attr_name] = attr
        return gold_candidate

    def _raise_error_on_internal_conflict(self, sen_id, gold_candidate):
        if self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_EXISTS]:
            current_gold = self.data[DocumentFields.SENS][sen_id][SenFields.GOLD_ATTRIBUTES]
            if set(gold_candidate.keys()) != set(current_gold.keys()):
                logging.error(
                    f"Conflict within already gold sentence {sen_id}:\nCurrent ({sen_id}):\n{current_gold}\nNew:\n{gold_candidate}")
                raise ValueError("Gold conflict")

    def _raise_error_on_external_conflict(self, sen_id, gold_candidate):
        text = self.data[DocumentFields.SENS][sen_id][SenFields.TEXT]
        current_gold = self.sen_to_gold_attrs.get_attrs(text)
        if current_gold is not None:
            current_gold_sens = self.sen_to_gold_attrs.get_sens(text)
            if set(gold_candidate.keys()) != set(current_gold.keys()):
                logging.error(
                    f"Conflict with current gold value of {sen_id}:\nCurrent ({current_gold_sens}):\n{current_gold}\nNew:\n{gold_candidate}")
                raise ValueError("Gold conflict")

    def _generate_gold_attrs(self, attributes):
        for attr in attributes:
            if attr["gold"]:
                yield attr["name"], {
                    AttributeFields.NAME: attr["name"],
                    AttributeFields.TYPE: None,
                    AttributeFields.VALUE: None,
                }

    def _generate_attributes(self, review_sheet, row_id):
        for col in range(ATTRIBUTE_OFFSET, review_sheet.max_column, ATTRIBUTE_STEP):
            label = review_sheet.cell(
                row=row_id, column=col+LABEL_OFFSET).value
            if label is None:
                continue
            review = review_sheet.cell(
                row=row_id, column=col+ATTRIBUTE_REVIEW_OFFSET).value
            if review is None:
                raise ValueError(
                    "Review field is not filled out for row " + str(row_id))
            if label not in ATTRIBUTES_TO_IGNORE:
                yield {
                    "name": label,
                    "gold": self._is_gold(review),
                }

    def _is_gold(self, review):
        return review == "OK" or review == "MISSING"

    def _is_error(self, attributes):
        return ERROR_LABEL in set([attr["name"] for attr in attributes])


def get_args():
    parser = argparse.ArgumentParser(description="")
    parser.add_argument("-r", "--review", default=None)
    parser.add_argument("-d", "--data-file", default=None)
    parser.add_argument("-g", "--gold-folder", default=None)
    parser.set_defaults(input_format="XLSX", output_format="JSON")
    return parser.parse_args()


def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s : " +
               "%(module)s (%(lineno)s) - %(levelname)s - %(message)s")
    args = get_args()
    converter = ReviewConverter(args.data_file, args.gold_folder)
    converter.convert(args.review)


if __name__ == "__main__":
    main()
