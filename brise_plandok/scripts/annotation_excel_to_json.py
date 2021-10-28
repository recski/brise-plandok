import sys
from brise_plandok.constants import SenFields
from brise_plandok.attrs_from_gold import SenToAttrMap, attrs_from_gold_sen
import json
import os
from tqdm import tqdm
import openpyxl

DATASET_SHEET_NAME = "Dataset"
FIRST_DATA_ROW = 3
ATTRIBUTE_OFFSET = 2
ATTRIBUTE_STEP = 2
LABEL_OFFSET = 1

ID_COL = 1
TEXT_COL = 2

ATTRIBUTES_TO_IGNORE = "DON'T ANNOTATE THIS SENTENCE"

JSON_ATTR = "/home/eszter/ownCloud/Shared/BRISE/data/plandok/ana/2021_09/json_attr"
DATA = "/home/eszter/research/brise-nlp/annotation/2021_09/full_data"
OLD_GOLD_FOLDER = "/home/eszter/research/brise-nlp/annotation/2021_09/gold"
XLSX_FOLDER = "/home/eszter/ownCloud/Shared/BRISE/data/plandok/ana/2021_09/xlsx"

GOLD_COLOR = "00FFD700"

class FullDataConverter:

    def __init__(self):
        self.sen_to_gold_attrs = SenToAttrMap(
            gold_dir=OLD_GOLD_FOLDER, fuzzy=True)

    def gen_sens_from_xlsx(self, fn):
        fn = os.path.join(XLSX_FOLDER, fn)
        fn = fn[:-1]
        wb = openpyxl.load_workbook(fn)
        sheet = wb[DATASET_SHEET_NAME]
        doc_id = os.path.basename(fn).split(".")[0]
        doc = {
            "id": doc_id,
            "sens": {},
        }
        for id, sen in self._gen_sens(sheet, self._get_json_attr(doc_id)):
            assert id not in doc["sens"]
            doc["sens"][id] = sen
        self._save(doc_id, doc)

    def _get_json_attr(self, doc_id):
        with open(os.path.join(JSON_ATTR, doc_id+".jsonl")) as f:
            return json.load(f)

    def _gen_sens(self, sheet, json_attr):
        for row in range(FIRST_DATA_ROW, sheet.max_row):
            id = sheet.cell(row=row, column=ID_COL).value
            text = sheet.cell(row=row, column=TEXT_COL).value
            gen_attributes = self._get_gen_attributes(json_attr, id)
            already_gold_on_annotation = self._is_already_gold(sheet, row)
            labels_gold_exists, gold_attr = self._get_gold_attr(text)
            yield id, {
                SenFields.ID: id,
                SenFields.TEXT: text,
                SenFields.MODALITY: None,
                SenFields.ALREADY_GOLD_ON_ANNOTATION: already_gold_on_annotation,
                SenFields.LABELS_GOLD_EXISTS: labels_gold_exists,
                SenFields.GOLD_ATTRIBUTES: gold_attr,
                SenFields.GEN_ATTRIBUTES_ON_ANNOTATION: gen_attributes,
                SenFields.ANNOTATED_ATTRIBUTES: {},
                SenFields.GEN_ATTRIBUTES: {},
                SenFields.SEGMENTATION_ERROR: False,
            }

    def _get_gen_attributes(self, doc, id):
        for section in doc["sections"]:
            for sen in section["sens"]:
                if sen["sen_id"] == id:
                    return sen[SenFields.GEN_ATTRIBUTES]

    def _is_already_gold(self, sheet, row):
        cell = sheet.cell(row=row, column=ID_COL)
        return cell.fill.fgColor.rgb == GOLD_COLOR

    def _save(self, doc_id, doc):
        data_file = os.path.join(DATA, doc_id + ".json")
        with open(data_file, "w") as f:
            json.dump(doc, f)

    def _get_gold_attr(self, text):
        sen = {SenFields.TEXT: text}
        attrs_from_gold_sen(sen, self.sen_to_gold_attrs, False)
        gold_attr = sen[SenFields.GOLD_ATTRIBUTES] if SenFields.GOLD_ATTRIBUTES in sen else []
        return sen[SenFields.LABELS_GOLD_EXISTS], gold_attr


def main():
    converter = FullDataConverter()
    for line in tqdm(sys.stdin):
        converter.gen_sens_from_xlsx(line)


if __name__ == "__main__":
    main()
