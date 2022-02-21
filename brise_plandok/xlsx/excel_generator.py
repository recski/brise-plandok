from brise_plandok.constants import GOLD_COLOR, GRAY_COLOR, ROW_HEIGHT, AttributeFields
from brise_plandok.annotation.attributes import ATTR_TO_CAT
from brise_plandok.constants import SenFields as SF
from brise_plandok.constants import DocumentFields as DF
from openpyxl.styles.fills import PatternFill
import logging
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation
from brise_plandok.data_utils import normalize_attribute_name


class ExcelGenerator:
    def __init__(self, output_file, CONSTANTS, doc, sen_to_gold_attrs=None):
        self.output_file = output_file
        self.sen_to_gold_attrs = sen_to_gold_attrs
        self.CONSTANTS = CONSTANTS
        self.doc = doc

    def generate_excel(self):
        workbook = openpyxl.load_workbook(self.input_template)
        self._fill_workbook(workbook)
        self._save_workbook(workbook)

    def _fill_workbook(self, workbook):
        main_sheet = workbook[self.CONSTANTS.MAIN_SHEET_NAME]
        row = self.CONSTANTS.FIRST_DATA_ROW
        self._modify_header(main_sheet)
        for sen_id, sen in self.doc[DF.SENS].items():
            self._fill_sentences(sen_id, main_sheet, row, sen)
            self._fill_attributes(sen, main_sheet, row)
            row += 1
        self._add_validation(main_sheet)
        self._set_row_height(main_sheet)

    def _fill_sentences(self, sen_id, main_sheet, row, sen):
        main_sheet.cell(row=row, column=self.CONSTANTS.SEN_ID_COL).value = sen_id
        main_sheet.cell(row=row, column=self.CONSTANTS.SEN_TEXT_COL).value = sen[SF.TEXT]
        if self._full_gold_exists(sen):
            self._color_gold(main_sheet, row, self.CONSTANTS.SEN_ID_COL)
            self._color_gold(main_sheet, row, self.CONSTANTS.SEN_TEXT_COL)
        main_sheet.cell(row=row, column=self.CONSTANTS.SEN_TEXT_COL).alignment = Alignment(
            wrapText=True
        )
        main_sheet.cell(row=row, column=self.CONSTANTS.SEN_TEXT_COL).font = Font(size=12)

    def _fill_attributes(self, sen, sheet, row):
        col = self.CONSTANTS.ATTRIBUTE_OFFSET
        self._fill_modality(sen, sheet, row)
        for attribute in self._gen_attributes(sen):
            attribute_name = normalize_attribute_name(attribute[AttributeFields.NAME])
            if attribute_name not in ATTR_TO_CAT:
                logging.warning(
                    f'"{attribute_name}" does not belong to any category - will be ignored'
                )
            else:
                self._fill_attribute(attribute, sen, sheet, col, row)
                col += self.CONSTANTS.ATTRIBUTE_STEP

    def _modify_header(self, sheet):
        raise NotImplementedError()

    def _add_validation(self, main_sheet):
        raise NotImplementedError()

    def _gen_attributes(self, sen):
        raise NotImplementedError()

    def _fill_modality(self, sen, sheet, row):
        raise NotImplementedError()

    def _fill_attribute(self, attribute, sen, sheet, col, row):
        raise NotImplementedError()

    def _labels_gold_exists(self, sen):
        return sen[SF.LABELS_GOLD_EXISTS]

    def _full_gold_exists(self, sen):
        return sen[SF.FULL_GOLD_EXISTS]

    def _color_gray(self, main_sheet, row, col):
        self._color(main_sheet, row, col, GRAY_COLOR)

    def _color_gold(self, main_sheet, row, col):
        self._color(main_sheet, row, col, GOLD_COLOR)

    def _color(self, main_sheet, row, col, color):
        main_sheet.cell(row=row, column=col).fill = PatternFill(fgColor=color, fill_type="solid")
        main_sheet.cell(row=row, column=col).font = Font(size=12)

    def _add_validations_for_attribute(self, main_sheet, row, col, category_val):
        category_val.add(main_sheet.cell(row=row, column=col))
        sub_data_val = DataValidation(
            type="list",
            formula1="==INDIRECT(${0}${1})".format(get_column_letter(col), str(row)),
        )
        main_sheet.add_data_validation(sub_data_val)
        sub_data_val.add(main_sheet.cell(row=row, column=col + self.CONSTANTS.LABEL_OFFSET))

    def _is_category_cell(self, col):
        return (col - self.CONSTANTS.ATTRIBUTE_OFFSET) % self.CONSTANTS.ATTRIBUTE_STEP == 0

    def _set_row_height(self, main_sheet):
        for row in range(self.CONSTANTS.FIRST_DATA_ROW, main_sheet.max_row + 1):
            main_sheet.row_dimensions[row].height = ROW_HEIGHT

    def _save_workbook(self, workbook):
        workbook.save(self.output_file)
        logging.info(f"DONE. Review xlsx was created to: {self.output_file}")
